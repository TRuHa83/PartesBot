from datetime import datetime, timedelta
from openpyxl import load_workbook

import shutil
import os


def get_filename(name):
    date = datetime.today()
    monday = date - timedelta(days=date.weekday())
    filename = f'{monday.strftime('%Y-%m-%d')}-{name[0][0].lower()}{name[1].lower()}.xlsx'
    return filename


def make_file(work_folder, destination, full_name):
    source = f'{work_folder}/partes.xlsx'
    shutil.copy2(source, destination)

    wb = load_workbook(destination)
    ws = wb.active

    worker = ws['A2'].value
    ws['A2'] = f'{worker} {full_name[0]} {full_name[1]}'

    today = datetime.today()
    monday = today.day - today.weekday()
    friday = monday + 4

    months = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio',
               'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

    ws['B4'] = months[today.month - 1]
    ws['B5'] = f'{monday}    al    {friday}'

    wb.save(destination)
    wb.close()


def prepare_file(data):
    full_name = data['name'].split(" ")
    data_folder = data['data_folder']

    filename = get_filename(full_name)

    destination = f'{data_folder}/{filename}'
    if not os.path.isfile(destination):
        make_file(data_folder, destination, full_name)

    return destination


def list_cells(cells):
    cells = cells.split(':')

    start_colm = cells[0][0]
    end_colm = cells[1][0]

    start_row = int(cells[0][1:])
    end_row = int(cells[1][1:])

    columns = [chr(col) for col in range(ord(start_colm), ord(end_colm) + 1)]
    column_pairs = [columns[i:i + 2] for i in range(0, len(columns), 2)]

    row = [row for row in range(start_row, end_row + 1)]
    row_pairs = [row[i:i + 2] for i in range(0, len(row), 2)]

    return column_pairs, row_pairs


# Funcion para convertir el tiempo de INTtoSTR o STRtoIN
def get_time(value):
    if isinstance(value, int):
        hour = str(value // 60)
        min = str(value % 60)

        if min == '0':
            min += '0'

        time = f'{hour}:{min}'

    elif value == '':
        time = value

    else:
        value = value.split(':')
        time = int(value[0]) * 60 + int(value[1])

    return time


def write_data(destination, data):
    wb = load_workbook(destination)
    ws = wb.active

    range_cell = 'A7:J18'
    column, row = list_cells(range_cell)
    position = datetime.today().weekday()

    daily_registry = data['daily_registry']

    for d in range(0, len(daily_registry)):
        site = daily_registry[d]['site']
        hour_entry = get_time(daily_registry[d]['entry'])
        hour_exit = get_time(daily_registry[d]['exit'])

        ws[f'{column[position][1]}{row[d][0]}'] = site
        ws[f'{column[position][0]}{row[d][0]}'] = hour_entry
        ws[f'{column[position][0]}{row[d][1]}'] = hour_exit

    wb.save(destination)
    wb.close()


def save(data):
    destination = prepare_file(data)
    write_data(destination, data)
